import io
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import load_workbook, Workbook
from app.db.database import get_db
from app.repositories.students import StudentRepository
from app.repositories.classes import ClassRepository

router = APIRouter(prefix="/students", tags=["students"])

REQUIRED_COLUMNS = ["姓名", "学号"]
OPTIONAL_COLUMNS = ["班级", "手机", "风险等级"]
COLUMN_MAP = {"姓名": "name", "学号": "student_id", "班级": "class_name", "手机": "phone", "风险等级": "risk_level"}


@router.post("/import")
async def import_students(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="文件为空或只有表头")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    for col in REQUIRED_COLUMNS:
        if col not in headers:
            raise HTTPException(status_code=400, detail=f"缺少必填列：{col}")

    col_index = {}
    for i, h in enumerate(headers):
        if h in COLUMN_MAP:
            col_index[COLUMN_MAP[h]] = i

    student_repo = StudentRepository(db)
    class_repo = ClassRepository(db)
    class_cache = {c.name: c.id for c in class_repo.list_all()}

    created = 0
    skipped = 0
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        name = str(row[col_index["name"]]).strip() if "name" in col_index and row[col_index["name"]] else ""
        student_id = str(row[col_index["student_id"]]).strip() if "student_id" in col_index and row[col_index["student_id"]] else ""

        if not name or not student_id:
            errors.append({"row": row_num, "error": "姓名或学号为空"})
            skipped += 1
            continue

        phone = str(row[col_index["phone"]]).strip() if "phone" in col_index and row[col_index["phone"]] else ""
        risk_level = str(row[col_index["risk_level"]]).strip() if "risk_level" in col_index and row[col_index["risk_level"]] else "low"
        if risk_level not in ("low", "medium", "high"):
            risk_level = "low"

        class_id = None
        if "class_name" in col_index:
            class_name = str(row[col_index["class_name"]]).strip() if row[col_index["class_name"]] else ""
            if class_name:
                if class_name not in class_cache:
                    new_class = class_repo.create(name=class_name)
                    class_cache[class_name] = new_class.id
                class_id = class_cache[class_name]

        try:
            student_repo.create(name=name, student_id=student_id, class_id=class_id, phone=phone, risk_level=risk_level)
            created += 1
        except Exception as e:
            errors.append({"row": row_num, "error": str(e)})
            skipped += 1

    return {"created": created, "skipped": skipped, "errors": errors, "total": len(rows) - 1}


@router.get("/import/template")
def download_template():
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "学生导入模板"

    # 表头
    headers = ["姓名", "学号", "班级", "手机", "风险等级"]
    ws.append(headers)

    # 示例数据
    ws.append(["李明", "2024001", "2024级软件1班", "13800001111", "low"])
    ws.append(["王红", "2024002", "2024级软件1班", "13800002222", "medium"])

    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12

    # 风险等级添加数据验证（下拉选项）
    dv = DataValidation(type="list", formula1='"low,medium,high"', allow_blank=True)
    dv.prompt = "请选择风险等级"
    dv.promptTitle = "风险等级"
    ws.add_data_validation(dv)
    dv.add('E2:E1000')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_import_template.xlsx"}
    )


@router.get("/export")
def export_students(class_id: str | None = Query(None), db: Session = Depends(get_db)):
    students = StudentRepository(db).list_all(class_id=class_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "学生数据"

    headers = ["姓名", "学号", "班级", "手机", "风险等级"]
    ws.append(headers)

    class_repo = ClassRepository(db)
    class_map = {c.id: c.name for c in class_repo.list_all()}
    risk_labels = {"low": "低", "medium": "中", "high": "高"}

    for s in students:
        ws.append([
            s.name,
            s.student_id,
            class_map.get(s.class_id, "") if s.class_id else "",
            s.phone,
            risk_labels.get(s.risk_level, s.risk_level),
        ])

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students_export.xlsx"}
    )

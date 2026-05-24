import io, json, logging
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from app.db.database import get_db
from app.services.ai.client import AIService

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/analysis", tags=["analysis"])


@router.post("/upload")
async def analyze_academic(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="文件为空或只有表头")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append(dict(zip(headers, [str(c) if c is not None else "" for c in row])))

    data_text = json.dumps(data_rows, ensure_ascii=False, indent=2)

    system_prompt = """你是一位高校学情分析专家。根据提供的学生成绩和考勤数据，生成学情分析报告。

请以 JSON 格式输出，字段如下：
- "class_overview": 班级整体情况总结（100-200字）
- "abnormal_students": 异常学生列表，每项含 name(姓名)、issue(问题描述)、severity(严重程度: high/medium/low)
- "academic_advice": 学风建设建议（3-5条）
- "grade_warnings": 学业预警学生列表，每项含 name(姓名)、course(课程)、risk(风险描述)
- "summary": 一句话总结（20字以内）

异常判定标准：
- 成绩 < 60 分 → 高风险
- 成绩 60-70 分 → 中风险
- 缺勤 > 3 次 → 需关注
- 请假 > 5 次 → 需关注"""

    user_message = f"以下是一个班级的学生成绩和考勤数据，请分析：\n\n{data_text}"

    ai = AIService()
    result = ai.chat(system_prompt=system_prompt, user_message=user_message)

    if result.get("success"):
        try:
            analysis = json.loads(result["content"].strip().removeprefix("```json").removesuffix("```").strip())
        except json.JSONDecodeError:
            analysis = {"raw": result["content"], "error": "AI 输出无法解析为 JSON"}
        return {"success": True, "analysis": analysis, "model": result["model"], "token_usage": result["token_usage"]}
    else:
        return {"success": False, "error": result.get("error", "分析失败")}
